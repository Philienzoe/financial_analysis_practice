"""
yf_field_scraper.py
───────────────────
Scrapes all field names from Yahoo Finance financial statements
(Balance Sheet, Income Statement, Cash Flow) for S&P 500 companies.
Captures repeated and non-repeated field names across all companies.

Run: python yf_field_scraper.py
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import yfinance as yf
import pandas as pd
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# ── Config ────────────────────────────────────────────────────────────
TICKER_FILE  = r"C:\Users\LENOVO\OneDrive\Desktop\financial_modeling\sp500_tickers.xlsx"   # from previous scrape
OUTPUT_FILE  = "../yf_fields_analysis.xlsx"
MAX_TICKERS  = None   # Set to e.g. 50 to test; None = all 503


# ═════════════════════════════════════════════════════════════════════
# STEP 1 — Collect all field names
# ═════════════════════════════════════════════════════════════════════

def get_fields(ticker: str) -> dict:
    """Fetch all row/field names from each financial statement."""
    result = {"balance_sheet": [], "income_statement": [], "cash_flow": []}
    try:
        t = yf.Ticker(ticker)
        for key, attr in [
            ("balance_sheet",    t.balance_sheet),
            ("income_statement", t.financials),
            ("cash_flow",        t.cashflow),
        ]:
            if attr is not None and not attr.empty:
                result[key] = list(attr.index)
    except Exception as e:
        print(f"    ⚠️  {ticker}: {e}")
    return result


def collect_all_fields(tickers: list) -> dict:
    """
    Returns dict with structure:
    {
      "balance_sheet":    Counter({field: count, ...}),
      "income_statement": Counter({field: count, ...}),
      "cash_flow":        Counter({field: count, ...}),
    }
    Also tracks which tickers have each field.
    """
    counters  = {
        "balance_sheet":    Counter(),
        "income_statement": Counter(),
        "cash_flow":        Counter(),
    }
    ticker_map = {
        "balance_sheet":    {},
        "income_statement": {},
        "cash_flow":        {},
    }

    total = len(tickers)
    for i, ticker in enumerate(tickers, 1):
        print(f"  [{i:03d}/{total}] {ticker:<8}", end="\r")
        fields = get_fields(ticker)

        for stmt in counters:
            for field in fields[stmt]:
                counters[stmt][field]  += 1
                if field not in ticker_map[stmt]:
                    ticker_map[stmt][field] = []
                ticker_map[stmt][field].append(ticker)

    print(f"\n  ✅ Done collecting fields from {total} tickers\n")
    return counters, ticker_map


# ═════════════════════════════════════════════════════════════════════
# STEP 2 — Analyse & Export
# ═════════════════════════════════════════════════════════════════════

def build_analysis_df(counter: Counter, ticker_map: dict, total: int) -> pd.DataFrame:
    """Build a DataFrame with field stats."""
    rows = []
    for field, count in counter.most_common():
        pct = round(count / total * 100, 1)
        rows.append({
            "field_name":     field,
            "count":          count,
            "pct_companies":  pct,
            "frequency":      "Universal (>90%)"  if pct >= 90
                         else "Common (50-90%)"   if pct >= 50
                         else "Occasional (10-50%)" if pct >= 10
                         else "Rare (<10%)",
            "repeated":       "YES" if count > 1 else "NO",
            "example_tickers": ", ".join(ticker_map[field][:5]),  # first 5
        })
    return pd.DataFrame(rows)


def export_to_excel(results: dict, ticker_map: dict, total: int, output_path: str):
    """Export full analysis to Excel with colored frequency bands."""

    # Color map for frequency
    COLORS = {
        "Universal (>90%)":      "C6EFCE",   # green
        "Common (50-90%)":       "FFEB9C",   # yellow
        "Occasional (10-50%)":   "FFCC99",   # orange
        "Rare (<10%)":           "FFC7CE",   # red
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

        summary_rows = []

        for stmt, counter in results.items():
            df = build_analysis_df(counter, ticker_map[stmt], total)
            df.to_excel(writer, sheet_name=stmt[:31], index=False)

            # Apply color formatting
            ws = writer.sheets[stmt[:31]]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                freq_cell = row[3]   # "frequency" column (index 3)
                freq_val  = freq_cell.value
                color     = COLORS.get(freq_val, "FFFFFF")
                fill      = PatternFill("solid", fgColor=color)
                for cell in row:
                    cell.fill = fill

            # Bold header
            for cell in ws[1]:
                cell.font = Font(bold=True)

            # Summary stats
            total_fields    = len(counter)
            universal       = sum(1 for f, c in counter.items() if c / total * 100 >= 90)
            common          = sum(1 for f, c in counter.items() if 50 <= c / total * 100 < 90)
            occasional      = sum(1 for f, c in counter.items() if 10 <= c / total * 100 < 50)
            rare            = sum(1 for f, c in counter.items() if c / total * 100 < 10)
            repeated        = sum(1 for f, c in counter.items() if c > 1)
            non_repeated    = sum(1 for f, c in counter.items() if c == 1)

            summary_rows.append({
                "statement":        stmt,
                "total_fields":     total_fields,
                "universal_>90%":   universal,
                "common_50-90%":    common,
                "occasional_10-50%":occasional,
                "rare_<10%":        rare,
                "repeated":         repeated,
                "non_repeated":     non_repeated,
            })

        # Summary sheet
        df_summary = pd.DataFrame(summary_rows)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        ws_sum = writer.sheets["Summary"]
        for cell in ws_sum[1]:
            cell.font = Font(bold=True)

        # Repeated fields sheet (appear in >1 company)
        all_repeated = []
        for stmt, counter in results.items():
            for field, count in counter.items():
                if count > 1:
                    all_repeated.append({
                        "statement": stmt,
                        "field_name": field,
                        "count": count,
                        "pct_companies": round(count / total * 100, 1),
                    })
        pd.DataFrame(all_repeated).sort_values(
            ["statement", "count"], ascending=[True, False]
        ).to_excel(writer, sheet_name="Repeated Fields", index=False)

        # Non-repeated fields sheet (unique to 1 company only)
        all_unique = []
        for stmt, counter in results.items():
            for field, count in counter.items():
                if count == 1:
                    all_unique.append({
                        "statement":  stmt,
                        "field_name": field,
                        "only_ticker": ticker_map[stmt][field][0],
                    })
        pd.DataFrame(all_unique).sort_values("statement").to_excel(
            writer, sheet_name="Non-Repeated Fields", index=False
        )

    print(f"  ✅ Results saved → {os.path.abspath(output_path)}\n")


# ═════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════

def main():
    # Load tickers
    if not os.path.exists(TICKER_FILE):
        print(f"  ❌ Ticker file not found: {TICKER_FILE}")
        print(f"     Run sp500_scraper.py --scrape first.\n")
        return

    df_tickers = pd.read_excel(TICKER_FILE)
    tickers    = df_tickers["ticker"].tolist()

    if MAX_TICKERS:
        tickers = tickers[:MAX_TICKERS]

    print(f"\n{'─'*60}")
    print(f"  Yahoo Finance Field Scraper")
    print(f"  Tickers: {len(tickers)}  |  Statements: 3")
    print(f"{'─'*60}\n")

    # Collect fields
    counters, ticker_map = collect_all_fields(tickers)

    # Print quick summary
    print(f"\n{'─'*60}")
    print(f"  FIELD COUNT SUMMARY")
    print(f"{'─'*60}")
    for stmt, counter in counters.items():
        repeated     = sum(1 for c in counter.values() if c > 1)
        non_repeated = sum(1 for c in counter.values() if c == 1)
        print(f"  {stmt:<20} total={len(counter):>4}  "
              f"repeated={repeated:>4}  unique={non_repeated:>4}")
    print(f"{'─'*60}\n")

    # Export
    export_to_excel(counters, ticker_map, len(tickers), OUTPUT_FILE)


if __name__ == "__main__":
    main()